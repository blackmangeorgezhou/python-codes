
import random
import time

import requests
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup

def get_json(url, page, lang_name):
  headers = {
    'Host': 'www.lagou.com',
    'Connection': 'keep-alive',
    'Content-Length': '23',
    'Origin': 'https://www.lagou.com',
    'X-Anit-Forge-Code': '0',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:61.0) Gecko/20100101 Firefox/61.0',
    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
    'Accept': 'application/json, text/javascript, */*; q=0.01',
    'X-Requested-With': 'XMLHttpRequest',
    'X-Anit-Forge-Token': 'None',
    'Referer': 'https://www.lagou.com/jobs/list_python?city=%E5%85%A8%E5%9B%BD&cl=false&fromSearch=true&labelWords=&suginput=',
    'Accept-Encoding': 'gzip, deflate, br',
    'Accept-Language': 'en-US,en;q=0.9,zh-CN;q=0.8,zh;q=0.7'
  }
  data = {'first': 'false', 'pn': page, 'kd': lang_name}
  json = requests.post(url, data, headers=headers).json()
  print(json)
  list_con = json['content']['positionResult']['result']
  info_list = []
  for i in list_con:
    info = []
    info.append(i.get('companyShortName', '无'))
    info.append(i.get('companyFullName', '无'))
    info.append(i.get('industryField', '无'))
    info.append(i.get('companySize', '无'))
    info.append(i.get('salary', '无'))
    info.append(i.get('city', '无'))
    info.append(i.get('education', '无'))
    info_list.append(info)
  return info_list

def fetchData():
  page = 1
  lang_name = 'python'
  wb = Workbook()

  ws1 = wb.active
  ws1.title = lang_name + '岗位爬虫数据'
  info = get_json('https://www.lagou.com/jobs/positionAjax.json?city=北京&needAddtionalResult=false', page, lang_name)
  if (len(info) > 0):
    for row in info:
      ws1.append(row)
    wb.save('{}岗位信息.xlsx'.format(lang_name))

  # for i in ['北京']:
  #     page = 1
  #     ws1 = wb.active
  #     ws1.title = lang_name
  #     url = 'https://www.lagou.com/jobs/positionAjax.json?city={}&needAddtionalResult=false'.format(i)
  #     while page < 2:
  #         info = get_json(url, page, lang_name)
  #         page += 1
  #         import time
  #         a = random.randint(10, 20)
  #         time.sleep(a)
  #         for row in info:
  #             ws1.append(row)
  # wb.save('{}职位信息.xlsx'.format(lang_name))

'''
from selenium import webdriver
driver = webdriver.Chrome()
driver.get('https://www.baidu.com/')
'''

def say(msg):
  print('我是{}'.format(msg))

def writeXls():
  dataList = []
  for i in range(10):
    data = {'name': 'tom{}'.format(i + 1) ,'age': 20 + i}
    dataList.append(data)
  print('====================')
  wb = Workbook()
  wbFile = wb.active
  wbFile.title = '数据'
  wbFile.append(['姓名', '年级'])

  for row in dataList:
    wbFile.append([row.get('name', '无'), row.get('age', '无')])
  wb.save('学员信息.xlsx')

def readXls():
  wb = load_workbook('学员信息.xlsx')
  for row in wb.active.values:
    for value in row:
      print(value,end='')
    print('')

def mockRequest():
  headers = {
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
    'Accept-Encoding': 'gzip, deflate',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Connection': 'keep-alive',
    'Host': 'souke.xdf.cn',
    'Pragma': 'no-cache',
    'Referer': 'http://souke.xdf.cn/search?cid=23&ccc=318&gc=0',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:61.0) Gecko/20100101 Firefox/61.0'
  }

  params = { 'cid': 23, 'ccc': 318, 'gc': 0, 'bdcode': '023002001' }

  result = requests.get('http://souke.xdf.cn/search', params, headers=headers)
  
  htmlSoup = BeautifulSoup(result.text, features='html.parser')
  # 格式化
  classDivList = htmlSoup.find_all('div', attrs={'class': 'm-courselist'})
  classList = []
  for div in classDivList:
    title = div.find('h2').find('a')
    clCode = title.attrs.get('id')
    clName = title.string
    classList.append({'clCode': clCode, 'clName': clName})
  wb = Workbook()
  ws = wb.active
  ws.append(['班号', '班级姓名'])

  for cl in classList:
    ws.append([cl.get('clCode'), cl.get('clName')])
  
  wb.save('新东方班级信息.xlsx')

  
# 当模块被直接执行时，将执行以下操作
if __name__ == '__main__':
  #writeXls()
  # readXls()
  mockRequest()